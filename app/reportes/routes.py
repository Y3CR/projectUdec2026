from flask import render_template, redirect, url_for, flash, request, send_file, current_app
from flask_login import login_required, current_user
from functools import wraps
from datetime import datetime, timedelta
from io import BytesIO
from app import db
from app.reportes import bp
from app.models import Solicitud, User, Espacio, Recurso


# ── Decoradores ────────────────────────────────────────────────────────────────

def operador_o_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        if current_user.get_role_name() not in ['operador', 'administrador']:
            flash('Acceso denegado.', 'danger')
            return redirect(url_for('admin.dashboard'))
        return f(*args, **kwargs)
    return decorated


def cualquier_usuario(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


# ── HU-18/19: Historial personal ──────────────────────────────────────────────

@bp.route('/historial')
@login_required
@cualquier_usuario
def historial():
    page = request.args.get('page', 1, type=int)
    estado_filtro = request.args.get('estado', '')
    tipo_filtro = request.args.get('tipo', '')
    fecha_desde_str = request.args.get('fecha_desde', '')
    fecha_hasta_str = request.args.get('fecha_hasta', '')

    # Admin y operador ven todo; docente/estudiante solo lo suyo
    role = current_user.get_role_name()
    if role in ['administrador', 'operador']:
        query = Solicitud.query
    else:
        query = Solicitud.query.filter_by(usuario_id=current_user.id)

    if estado_filtro:
        query = query.filter_by(estado=estado_filtro)
    if tipo_filtro:
        query = query.filter_by(tipo=tipo_filtro)
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d')
            query = query.filter(Solicitud.fecha_creacion >= fecha_desde)
        except ValueError:
            pass
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Solicitud.fecha_creacion < fecha_hasta)
        except ValueError:
            pass

    solicitudes = query.order_by(Solicitud.fecha_creacion.desc()).paginate(
        page=page, per_page=15, error_out=False
    )

    return render_template('reportes/historial.html',
                           solicitudes=solicitudes,
                           estado_filtro=estado_filtro,
                           tipo_filtro=tipo_filtro,
                           fecha_desde=fecha_desde_str,
                           fecha_hasta=fecha_hasta_str)


# ── HU-16/17: Reportes operador y admin ───────────────────────────────────────

@bp.route('/reportes')
@login_required
@operador_o_admin
def reportes():
    # Filtros
    estado_filtro = request.args.get('estado', '')
    tipo_filtro = request.args.get('tipo', '')
    usuario_filtro = request.args.get('usuario_id', 0, type=int)
    fecha_desde_str = request.args.get('fecha_desde', '')
    fecha_hasta_str = request.args.get('fecha_hasta', '')
    page = request.args.get('page', 1, type=int)

    query = Solicitud.query

    if estado_filtro:
        query = query.filter_by(estado=estado_filtro)
    if tipo_filtro:
        query = query.filter_by(tipo=tipo_filtro)
    if usuario_filtro:
        query = query.filter_by(usuario_id=usuario_filtro)
    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, '%Y-%m-%d')
            query = query.filter(Solicitud.fecha_creacion >= fecha_desde)
        except ValueError:
            pass
    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Solicitud.fecha_creacion < fecha_hasta)
        except ValueError:
            pass

    solicitudes = query.order_by(Solicitud.fecha_creacion.desc()).paginate(
        page=page, per_page=15, error_out=False
    )

    # Estadísticas resumen
    total = query.count() if hasattr(query, 'count') else 0
    stats = {
        'total': Solicitud.query.count(),
        'pendientes': Solicitud.query.filter_by(estado='pendiente').count(),
        'aprobadas': Solicitud.query.filter_by(estado='aprobada').count(),
        'rechazadas': Solicitud.query.filter_by(estado='rechazada').count(),
        'devueltas': Solicitud.query.filter_by(estado='devuelta').count(),
    }

    usuarios = User.query.filter(
        User.role.has(name='docente') | User.role.has(name='estudiante')
    ).order_by(User.nombre).all()

    return render_template('reportes/reportes.html',
                           solicitudes=solicitudes,
                           stats=stats,
                           usuarios=usuarios,
                           estado_filtro=estado_filtro,
                           tipo_filtro=tipo_filtro,
                           usuario_filtro=usuario_filtro,
                           fecha_desde=fecha_desde_str,
                           fecha_hasta=fecha_hasta_str)


# ── Exportar PDF ───────────────────────────────────────────────────────────────

@bp.route('/exportar/pdf')
@login_required
@operador_o_admin
def exportar_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    # Aplicar mismos filtros que la vista de reportes
    estado_filtro = request.args.get('estado', '')
    tipo_filtro = request.args.get('tipo', '')
    usuario_filtro = request.args.get('usuario_id', 0, type=int)
    fecha_desde_str = request.args.get('fecha_desde', '')
    fecha_hasta_str = request.args.get('fecha_hasta', '')

    query = Solicitud.query
    if estado_filtro:
        query = query.filter_by(estado=estado_filtro)
    if tipo_filtro:
        query = query.filter_by(tipo=tipo_filtro)
    if usuario_filtro:
        query = query.filter_by(usuario_id=usuario_filtro)
    if fecha_desde_str:
        try:
            query = query.filter(Solicitud.fecha_creacion >= datetime.strptime(fecha_desde_str, '%Y-%m-%d'))
        except ValueError:
            pass
    if fecha_hasta_str:
        try:
            query = query.filter(Solicitud.fecha_creacion < datetime.strptime(fecha_hasta_str, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass

    solicitudes = query.order_by(Solicitud.fecha_creacion.desc()).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    verde = colors.HexColor('#1a6b3a')
    elements = []

    # Título
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                 textColor=verde, fontSize=16, spaceAfter=6)
    elements.append(Paragraph('Reporte de Préstamos — UCundinamarca', title_style))
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9,
                               textColor=colors.grey, spaceAfter=12)
    elements.append(Paragraph(f'Generado el {datetime.now().strftime("%d/%m/%Y %H:%M")} | Total registros: {len(solicitudes)}', sub_style))
    elements.append(Spacer(1, 0.3*cm))

    # Tabla
    headers = ['#', 'Usuario', 'Rol', 'Tipo', 'Item', 'Fecha inicio', 'Fecha fin', 'Estado', 'Tiempo uso']
    data = [headers]

    for s in solicitudes:
        tiempo = ''
        if s.tiempo_uso_minutos is not None:
            tiempo = f"{s.tiempo_uso_minutos // 60}h {s.tiempo_uso_minutos % 60}min"
        data.append([
            str(s.id),
            f"{s.usuario.nombre} {s.usuario.apellido}",
            s.usuario.get_role_name().capitalize(),
            s.tipo.capitalize(),
            s.get_item_nombre(),
            s.fecha_inicio.strftime('%d/%m/%Y %H:%M'),
            s.fecha_fin.strftime('%d/%m/%Y %H:%M'),
            s.estado.capitalize(),
            tiempo or '—'
        ])

    col_widths = [1*cm, 4*cm, 2.5*cm, 2*cm, 4.5*cm, 3.5*cm, 3.5*cm, 2.5*cm, 2.5*cm]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), verde),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f7f3')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    filename = f"reporte_prestamos_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buffer, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


# ── Exportar Excel ─────────────────────────────────────────────────────────────

@bp.route('/exportar/excel')
@login_required
@operador_o_admin
def exportar_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    estado_filtro = request.args.get('estado', '')
    tipo_filtro = request.args.get('tipo', '')
    usuario_filtro = request.args.get('usuario_id', 0, type=int)
    fecha_desde_str = request.args.get('fecha_desde', '')
    fecha_hasta_str = request.args.get('fecha_hasta', '')

    query = Solicitud.query
    if estado_filtro:
        query = query.filter_by(estado=estado_filtro)
    if tipo_filtro:
        query = query.filter_by(tipo=tipo_filtro)
    if usuario_filtro:
        query = query.filter_by(usuario_id=usuario_filtro)
    if fecha_desde_str:
        try:
            query = query.filter(Solicitud.fecha_creacion >= datetime.strptime(fecha_desde_str, '%Y-%m-%d'))
        except ValueError:
            pass
    if fecha_hasta_str:
        try:
            query = query.filter(Solicitud.fecha_creacion < datetime.strptime(fecha_hasta_str, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass

    solicitudes = query.order_by(Solicitud.fecha_creacion.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Préstamos'

    # Estilos
    verde_fill = PatternFill(start_color='1a6b3a', end_color='1a6b3a', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    data_font = Font(name='Calibri', size=10)
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells('A1:I1')
    ws['A1'] = f'Reporte de Préstamos — UCundinamarca | Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A1'].font = Font(name='Calibri', bold=True, size=13, color='1a6b3a')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 25

    # Encabezados
    headers = ['#', 'Usuario', 'Rol', 'Tipo', 'Item solicitado',
               'Fecha inicio', 'Fecha fin', 'Estado', 'Tiempo de uso']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = verde_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[2].height = 20

    # Datos
    for row_num, s in enumerate(solicitudes, 3):
        tiempo = ''
        if s.tiempo_uso_minutos is not None:
            tiempo = f"{s.tiempo_uso_minutos // 60}h {s.tiempo_uso_minutos % 60}min"

        fila = [
            s.id,
            f"{s.usuario.nombre} {s.usuario.apellido}",
            s.usuario.get_role_name().capitalize(),
            s.tipo.capitalize(),
            s.get_item_nombre(),
            s.fecha_inicio.strftime('%d/%m/%Y %H:%M'),
            s.fecha_fin.strftime('%d/%m/%Y %H:%M'),
            s.estado.capitalize(),
            tiempo or '—'
        ]
        fill_color = 'f0f7f3' if row_num % 2 == 0 else 'FFFFFF'
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        for col_num, value in enumerate(fila, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = data_font
            cell.alignment = center
            cell.border = border
            cell.fill = row_fill

    # Ancho de columnas
    col_widths = [6, 25, 15, 12, 30, 20, 20, 15, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"reporte_prestamos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buffer,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)